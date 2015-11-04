import openpyxl as ox, MySQLdb as my, string

def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

wb = ox.load_workbook(filename='High_Value_Data_Sets.xlsx', use_iterators=True)
ws = wb.get_active_sheet()
conn = my.connect(host='localhost', user='root', passwd='ldhagen')
c1 = conn.cursor()
c1.execute(r'drop database if exists TxInmates;')
c1.execute(r'create database TxInmates;')
c1.execute(r'use TxInmates;')
c1.execute("""create table inmates (SID_NUMBER INT(8) NOT NULL, PRIMARY KEY(SID_NUMBER), TDCJ_NUMBER INT(8), NAME CHAR(80), CURRENT_FACILITY CHAR(80),
           GENDER CHAR(1), RACE CHAR(1), DOB DATE, PROJECTED_RELEASE DATE, MAXIMUM_SENTENCE_DATE DATE, PAROLE_ELIGIBILITY_DATE DATE, CASE_NUMBER CHAR(80), COUNTY CHAR(45),
           OFFENSE_CODE INT(30), OFFENSE CHAR(80), SENTENCE_DATE DATE, OFFENSE_DATE DATE, SENTENCE_YEARS CHAR(20),
           LAST_PAROLE_DECISION CHAR(80), NEXT_PAROLE_REVIEW_DATE DATE, PAROLE_REVIEW_STATUS CHAR(80))""")
count = 0
row_iter = iter_rows(ws)
row_iter.next() #skip header row
for x in row_iter:
    c1.execute(r'insert into inmates (SID_NUMBER, TDCJ_NUMBER, NAME, CURRENT_FACILITY, GENDER, RACE, DOB, PROJECTED_RELEASE, MAXIMUM_SENTENCE_DATE, PAROLE_ELIGIBILITY_DATE, CASE_NUMBER, COUNTY, OFFENSE_CODE, OFFENSE, SENTENCE_DATE, OFFENSE_DATE, SENTENCE_YEARS, LAST_PAROLE_DECISION, NEXT_PAROLE_REVIEW_DATE, PAROLE_REVIEW_STATUS) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(x))
    count += 1
    print count
conn.commit()

