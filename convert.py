#!/bin/env python

import openpyxl as ox, sqlite3 as sl, string

def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

wb = ox.load_workbook(filename='High_Value_Data_Sets.xlsx', use_iterators=True)
ws = wb.get_active_sheet()

conn = sl.connect('TexScript.db')
c1 = conn.cursor()
c1.execute('create table inmates ("SID Number" ,"TDCJ Number" ,"Name" ,"Current Facility" ,"Gender" ,"Race" ,"DOB" ,"Projected Release" ,"Maximum Sentence Date" ,"Parole Eligibility Date" ,"Case Number" ,"County" ,"Offense Code" ,"Offense" ,"Sentence Date" ,"Offense Date" ,"Sentence Years" ,"Last Parole Decision" ,"Next Parole Review Date" ,"Parole Review Status")')

count = 0
row_iter = iter_rows(ws)
row_iter.next() #skip header row
for x in row_iter:
    c1.execute('insert into inmates ("SID Number" ,"TDCJ Number" ,"Name" ,"Current Facility" ,"Gender" ,"Race" ,"DOB" ,"Projected Release" ,"Maximum Sentence Date" ,"Parole Eligibility Date" ,"Case Number" ,"County" ,"Offense Code" ,"Offense" ,"Sentence Date" ,"Offense Date" ,"Sentence Years" ,"Last Parole Decision" ,"Next Parole Review Date" ,"Parole Review Status") values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ? )', x)
    count += 1
    print count
conn.commit()
