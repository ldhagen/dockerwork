import openpyxl as ox, MySQLdb as my, string

def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

wb = ox.load_workbook(filename='peps300.xlsx', read_only=True)
ws = wb.active
conn = my.connect(host='localhost', user='root')
c1 = conn.cursor()
c1.execute(r'drop database if exists Schools;')
c1.execute(r'create database Schools;')
c1.execute(r'use Schools;')
c1.execute("""create table schools (OPEID CHAR(20), NAME CHAR(100), ADDRESS CHAR(100), CITY CHAR(100),
           STATE CHAR(20), STATE_DESC CHAR(80), ZIP CHAR(20), ZIP_EXT CHAR(20), PROGRAM_LENGTH INT(2),
           SCHOOL_TYPE INT(2), YEAR_1 INT(4), NUM_1 INT(10), DENOM_1 INT(10), DRATE_1 DECIMAL(5,2),
           PRATE_1 CHAR(20), ETHNIC_CODE CHAR(20), CONG_DIS CHAR(20), REGION CHAR(20), AVERAGE_OR_GREATER_THAN_30 CHAR(20),
           RATE_TYPE CHAR(1), YEAR_2 INT(4), NUM_2 INT(10), DENOM_2 INT(10), DRATE_2 DECIMAL(5,2), PRATE_2 CHAR(20),
           YEAR_3 INT(4), NUM_3 INT(10), DENOM_3 INT(10), DRATE_3 DECIMAL(5,2), PRATE_3 CHAR(20))""")
count = 0
row_iter = iter_rows(ws)
row_iter.next() #skip header row
for x in row_iter:
    c1.execute(r'insert into schools (OPEID, NAME, ADDRESS, CITY, STATE, STATE_DESC, ZIP, ZIP_EXT, PROGRAM_LENGTH, SCHOOL_TYPE, YEAR_1, NUM_1, DENOM_1, DRATE_1, PRATE_1, ETHNIC_CODE, CONG_DIS, REGION, AVERAGE_OR_GREATER_THAN_30,RATE_TYPE, YEAR_2, NUM_2, DENOM_2, DRATE_2, PRATE_2, YEAR_3, NUM_3, DENOM_3, DRATE_3, PRATE_3) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(x))
    count += 1
    print count
conn.commit()
